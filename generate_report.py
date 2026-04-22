"""
Generate an Excel report for the Gold forward-test portfolio.
Can be run at any time, but designed for end-of-month summary.

Usage:
    python generate_report.py
    python portfolio_tracker.py --report   (same thing)
"""

from __future__ import annotations

import json
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

BKK_TZ = timezone(timedelta(hours=7))
BASE_DIR = Path(__file__).parent


def load_data():
    pf_path = BASE_DIR / "portfolio_state.json"
    tl_path = BASE_DIR / "trade_log.json"

    if not pf_path.exists():
        raise FileNotFoundError(
            "ยังไม่มีข้อมูลพอร์ต — ต้องรัน portfolio_tracker.py อย่างน้อย 1 ครั้งก่อนสร้างรายงาน\n"
            "Run: uv run --python 3.12 --with pandas --with numpy --with requests --with yfinance --with openpyxl portfolio_tracker.py"
        )
    with open(pf_path) as f:
        portfolio = json.load(f)
    trades = []
    if tl_path.exists():
        with open(tl_path) as f:
            trades = json.load(f)
    return portfolio, trades


def generate_report(output_path: str | None = None):
    portfolio, trades = load_data()

    if output_path is None:
        output_path = str(BASE_DIR / "Gold_Forward_Test_Report.xlsx")

    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_properties.tabColor = "FFD700"

    # Styles
    title_font = Font(name="Arial", bold=True, size=16, color="1a1a1a")
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E50")
    gold_fill = PatternFill("solid", fgColor="FFF8DC")
    currency_fmt = '#,##0.00'
    pct_fmt = '0.00%'
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    label_font = Font(name="Arial", bold=True, size=11, color="2C3E50")
    value_font = Font(name="Arial", size=11)

    ws_sum.column_dimensions["A"].width = 5
    ws_sum.column_dimensions["B"].width = 28
    ws_sum.column_dimensions["C"].width = 22

    # Title
    ws_sum.merge_cells("B2:C2")
    c = ws_sum["B2"]
    c.value = "Gold Forward Test Report"
    c.font = title_font
    c.alignment = Alignment(horizontal="left")

    ws_sum.merge_cells("B3:C3")
    c = ws_sum["B3"]
    c.value = f"Period: {portfolio.get('test_start', '')[:10]} → {portfolio.get('test_end', '')[:10]}"
    c.font = Font(name="Arial", size=10, color="666666")

    # Summary table
    summary_data = [
        ("Initial Capital", portfolio.get("initial_cash", 100000)),
        ("Current Cash", portfolio.get("cash", 0)),
        ("Gold Holdings (baht-weight)", portfolio.get("gold_units", 0)),
        ("Last Price (THB/baht-wt)", portfolio.get("last_price", 0)),
    ]

    row = 5
    for label, val in summary_data:
        ws_sum.cell(row=row, column=2, value=label).font = label_font
        c = ws_sum.cell(row=row, column=3, value=val)
        c.font = value_font
        if isinstance(val, float) and "baht-weight" not in label:
            c.number_format = currency_fmt
        ws_sum.cell(row=row, column=2).border = thin_border
        ws_sum.cell(row=row, column=3).border = thin_border
        row += 1

    # Portfolio value formula
    row += 1
    ws_sum.cell(row=row, column=2, value="Portfolio Value").font = Font(
        name="Arial", bold=True, size=12, color="1a6b1a"
    )
    c = ws_sum.cell(row=row, column=3)
    # =cash + gold_units * last_price → C6 + C7 * C8
    c.value = "=C6+C7*C8"
    c.font = Font(name="Arial", bold=True, size=12, color="1a6b1a")
    c.number_format = currency_fmt
    c.border = thin_border
    portfolio_value_cell = f"C{row}"

    row += 1
    ws_sum.cell(row=row, column=2, value="P&L (THB)").font = label_font
    c = ws_sum.cell(row=row, column=3)
    c.value = f"={portfolio_value_cell}-C5"
    c.font = Font(name="Arial", bold=True, size=11)
    c.number_format = '#,##0.00;(#,##0.00);"-"'
    c.border = thin_border

    row += 1
    ws_sum.cell(row=row, column=2, value="Return (%)").font = label_font
    c = ws_sum.cell(row=row, column=3)
    c.value = f"=IF(C5=0,0,({portfolio_value_cell}-C5)/C5)"
    c.font = Font(name="Arial", bold=True, size=11)
    c.number_format = pct_fmt
    c.border = thin_border

    row += 1
    ws_sum.cell(row=row, column=2, value="Total Trades").font = label_font
    ws_sum.cell(row=row, column=3, value=len(trades)).font = value_font
    ws_sum.cell(row=row, column=2).border = thin_border
    ws_sum.cell(row=row, column=3).border = thin_border

    row += 1
    ws_sum.cell(row=row, column=2, value="Current Position").font = label_font
    ws_sum.cell(row=row, column=3, value=portfolio.get("position", "CASH")).font = value_font

    # ── Sheet 2: Trade Log ─────────────────────────────────────────────
    ws_trades = wb.create_sheet("Trade Log")
    ws_trades.sheet_properties.tabColor = "2C3E50"

    headers = [
        "No.", "Date/Time", "Action", "Price (THB)",
        "Units (baht-wt)", "Trade Value (THB)",
        "Cash Before", "Cash After",
        "Gold Before", "Gold After",
    ]
    col_widths = [6, 20, 10, 18, 16, 18, 18, 18, 16, 16]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws_trades.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
        ws_trades.column_dimensions[get_column_letter(col_idx)].width = w

    buy_fill = PatternFill("solid", fgColor="E8F5E9")
    sell_fill = PatternFill("solid", fgColor="FFEBEE")

    for i, t in enumerate(trades, 1):
        row = i + 1
        fill = buy_fill if t["action"] == "BUY" else sell_fill

        data = [
            i,
            t["timestamp"][:16].replace("T", " "),
            t["action"],
            t["price"],
            t["units"],
            t["value"],
            t.get("cash_before", ""),
            t.get("cash_after", ""),
            t.get("gold_before", ""),
            t.get("gold_after", ""),
        ]

        for col_idx, val in enumerate(data, 1):
            c = ws_trades.cell(row=row, column=col_idx, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = fill
            c.border = thin_border
            c.alignment = Alignment(horizontal="center")
            if col_idx in (4, 6, 7, 8):
                c.number_format = currency_fmt
            elif col_idx in (5, 9, 10):
                c.number_format = '#,##0.0000'

    # Action column: bold + color
    for i in range(len(trades)):
        row = i + 2
        c = ws_trades.cell(row=row, column=3)
        if c.value == "BUY":
            c.font = Font(name="Arial", bold=True, size=10, color="1a6b1a")
        else:
            c.font = Font(name="Arial", bold=True, size=10, color="CC0000")

    # ── Sheet 3: Portfolio Value Over Time (after each trade) ──────────
    if trades:
        ws_val = wb.create_sheet("Portfolio Value")
        ws_val.sheet_properties.tabColor = "27AE60"

        val_headers = ["Trade No.", "Date", "Action", "Portfolio Value (THB)"]
        val_widths = [12, 20, 10, 22]

        for col_idx, (h, w) in enumerate(zip(val_headers, val_widths), 1):
            c = ws_val.cell(row=1, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border
            ws_val.column_dimensions[get_column_letter(col_idx)].width = w

        # Row 2: starting point (before any trade)
        ws_val.cell(row=2, column=1, value=0).font = Font(name="Arial", size=10)
        ws_val.cell(row=2, column=2, value=portfolio.get("test_start", "")[:10]).font = Font(name="Arial", size=10)
        ws_val.cell(row=2, column=3, value="START").font = Font(name="Arial", bold=True, size=10, color="336699")
        c = ws_val.cell(row=2, column=4, value=portfolio.get("initial_cash", 100000))
        c.number_format = currency_fmt
        c.font = Font(name="Arial", size=10)

        for i, t in enumerate(trades, 1):
            row = i + 2
            ws_val.cell(row=row, column=1, value=i).font = Font(name="Arial", size=10)
            ws_val.cell(row=row, column=2, value=t["timestamp"][:10]).font = Font(name="Arial", size=10)
            action_cell = ws_val.cell(row=row, column=3, value=t["action"])
            if t["action"] == "BUY":
                action_cell.font = Font(name="Arial", bold=True, size=10, color="1a6b1a")
            else:
                action_cell.font = Font(name="Arial", bold=True, size=10, color="CC0000")

            # After a SELL trade, portfolio value = cash_after
            # After a BUY trade, portfolio value = gold_after * price
            if t["action"] == "SELL":
                val = t.get("cash_after", t["value"])
            else:
                val = t.get("gold_after", 0) * t["price"]

            c = ws_val.cell(row=row, column=4, value=round(val, 2))
            c.number_format = currency_fmt
            c.font = Font(name="Arial", size=10)

        # Line chart
        chart = LineChart()
        chart.title = "Portfolio Value Over Trades"
        chart.y_axis.title = "THB"
        chart.x_axis.title = "Trade"
        chart.style = 10
        chart.width = 22
        chart.height = 14

        data_ref = Reference(ws_val, min_col=4, min_row=1,
                             max_row=len(trades) + 2)
        cat_ref = Reference(ws_val, min_col=1, min_row=2,
                            max_row=len(trades) + 2)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.series[0].graphicalProperties.line.width = 25000

        ws_val.add_chart(chart, "F2")

    # ── Sheet 4: Daily Performance ────────────────────────────────────
    daily_history_path = BASE_DIR / "daily_history.json"
    daily_history = []
    if daily_history_path.exists():
        with open(daily_history_path) as f:
            daily_history = json.load(f)

    if daily_history:
        ws_daily = wb.create_sheet("Daily Performance")
        ws_daily.sheet_properties.tabColor = "3498DB"

        daily_headers = [
            "Day", "Date", "Gold Price", "Cash (THB)",
            "Gold (baht-wt)", "Gold Value (THB)",
            "Portfolio Value (THB)", "P&L (THB)", "P&L (%)",
            "Signal", "Buys", "Sells", "Total Trades",
        ]
        daily_widths = [6, 14, 16, 18, 14, 18, 20, 18, 10, 10, 8, 8, 12]

        for col_idx, (h, w) in enumerate(zip(daily_headers, daily_widths), 1):
            c = ws_daily.cell(row=1, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
            ws_daily.column_dimensions[get_column_letter(col_idx)].width = w

        green_font = Font(name="Arial", size=10, color="1a6b1a")
        red_font = Font(name="Arial", size=10, color="CC0000")
        normal_font = Font(name="Arial", size=10)
        even_fill = PatternFill("solid", fgColor="F7F9FC")

        for i, day in enumerate(daily_history):
            row = i + 2
            fill = even_fill if i % 2 == 0 else PatternFill()

            row_data = [
                i + 1,
                day.get("date", ""),
                day.get("price", 0),
                day.get("cash", 0),
                day.get("gold_units", 0),
                day.get("gold_value", 0),
                day.get("port_value", 0),
                day.get("pnl", 0),
                day.get("pnl_pct", 0) / 100 if day.get("pnl_pct", 0) != 0 else 0,
                day.get("signal", ""),
                day.get("buys_today", 0),
                day.get("sells_today", 0),
                day.get("total_trades", 0),
            ]

            for col_idx, val in enumerate(row_data, 1):
                c = ws_daily.cell(row=row, column=col_idx, value=val)
                c.font = normal_font
                c.fill = fill
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")

                # Number formats
                if col_idx in (3, 4, 6, 7):    # prices & currency
                    c.number_format = currency_fmt
                elif col_idx == 5:               # gold units
                    c.number_format = '#,##0.0000'
                elif col_idx == 8:               # P&L THB
                    c.number_format = '#,##0.00;(#,##0.00);"-"'
                    if isinstance(val, (int, float)):
                        c.font = green_font if val >= 0 else red_font
                elif col_idx == 9:               # P&L %
                    c.number_format = pct_fmt
                    if isinstance(val, (int, float)):
                        c.font = green_font if val >= 0 else red_font

        # Portfolio Value line chart
        if len(daily_history) >= 2:
            chart = LineChart()
            chart.title = "Daily Portfolio Value"
            chart.y_axis.title = "THB"
            chart.x_axis.title = "Date"
            chart.style = 10
            chart.width = 28
            chart.height = 14

            data_ref = Reference(ws_daily, min_col=7, min_row=1,
                                 max_row=len(daily_history) + 1)
            cat_ref = Reference(ws_daily, min_col=2, min_row=2,
                                max_row=len(daily_history) + 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cat_ref)
            chart.series[0].graphicalProperties.line.width = 25000

            ws_daily.add_chart(chart, "O2")

            # P&L % chart
            chart2 = LineChart()
            chart2.title = "Cumulative Return (%)"
            chart2.y_axis.title = "%"
            chart2.y_axis.numFmt = '0.00%'
            chart2.style = 10
            chart2.width = 28
            chart2.height = 14

            pnl_ref = Reference(ws_daily, min_col=9, min_row=1,
                                max_row=len(daily_history) + 1)
            chart2.add_data(pnl_ref, titles_from_data=True)
            chart2.set_categories(cat_ref)
            chart2.series[0].graphicalProperties.line.width = 25000

            ws_daily.add_chart(chart2, "O18")

    # ── Save ──────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"Report saved: {output_path}")
    return output_path


if __name__ == "__main__":
    generate_report()
